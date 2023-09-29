
#code written by YASHARTH DUBEY
#IIIT DWD 
import tkinter
from tkinter import IntVar
import matplotlib.pyplot as plt
from PIL import ImageTk,Image
import webbrowser
import bs4 
import requests
import openpyxl as op
import os
import xlsxwriter
import xlrd
import tkinter.tix
import tkinter.ttk
import numpy as np
import sys
#class to add the scroll bar in the frame
class Scrollable(tkinter.Frame):

    def __init__(self, frame, width=16):

        scrollbar = tkinter.Scrollbar(frame, width=width)
        scrollbar.pack(side=tkinter.RIGHT, fill=tkinter.Y, expand=False)

        self.canvas = tkinter.Canvas(frame, yscrollcommand=scrollbar.set)
        self.canvas.pack(side=tkinter.LEFT, fill=tkinter.BOTH, expand=True)

        scrollbar.config(command=self.canvas.yview)

        self.canvas.bind('<Configure>', self.__fill_canvas)

        # base class initialization
        tkinter.Frame.__init__(self, frame)         

        # assign this obj (the inner frame) to the windows item of the canvas
        self.windows_item = self.canvas.create_window(0,0, window=self, anchor=tkinter.NW)


    def __fill_canvas(self, event):

        canvas_width = event.width
        self.canvas.itemconfig(self.windows_item, width = canvas_width)        

    def update(self):

        self.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox(self.windows_item))
#try block
try:
    window = tkinter.Tk()
    window.title("COVID-19 TRACKER")
    window.geometry("400x250")
    canvas = tkinter.Canvas(window,width = 400 , height = 250)
    image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
    canvas.create_image( 199, 124 ,image = image)
    canvas.place(x = 0, y = 0)
    #this function gives us the data in graphical representation
    def graphs():
        window.destroy()
        window2 = tkinter.Tk()
        window2.title("COVID-19 TRACKER")
        window2.geometry("400x250")
        canvas = tkinter.Canvas(window2,width = 400 , height = 250)
        image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
        canvas.create_image( 199, 124 ,image = image)
        canvas.place(x = 0, y = 0)
        #graph of the death/confirmed cases per day
        def deathc():
            window2.destroy()
            window = tkinter.Tk()
            window.resizable(0,0)
            window.geometry("400x250")
            window.title("COVID-19 TRACKER")
            canvas = tkinter.Canvas(window,width = 400 , height = 250)
            image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
            canvas.create_image( 199, 124 ,image = image)
            canvas.place(x = 0, y = 0)
            tkinter.Label(window,text = "NUMBER OF COUNTRIES:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
            ent1 = tkinter.Entry(window,bd  = "2",bg = "skyblue2")
            ent1.place(x = 160 , y = 118)
            var1 = IntVar()
            var2 = IntVar()
            tkinter.Checkbutton(window,text = "BAR GRAPH",variable = var1,border = "0",bg = "skyblue2",onvalue = 1, offvalue = 0).place(x = 20, y = 148)
            tkinter.Checkbutton(window,text = "LINE GRAPH",variable = var2,border = "0",bg = "skyblue2",onvalue = 1,offvalue = 0).place(x = 150 , y = 148)
            tkinter.Label(window,text = "*WE RECOMMEND YOU TO SEE BAR GRAPH FOR ONE COUNTRY",bg = "skyblue2",border = "0").place(x = 20 , y = 180)
            #to show the graph 
            def showwwac():
                loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                loc1  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
                wb1 = xlrd.open_workbook(loc1)
                sheet1 = wb1.sheet_by_index(0)
                window.destroy()
                window2 = tkinter.Tk()
                window2.resizable(0,0)
                window2.geometry("400x250")
                window2.title("COVID-19 TRACKER")
                canvas1 = tkinter.Canvas(window2,width = 400 , height = 250)
                image1 = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
                canvas1.create_image( 199, 124 ,image = image1)
                canvas1.place(x = 0, y = 0)
                tkinter.Label(window2,text = "TYPE THE COUNTRIES NAME WITH A COMMA IN END", bg ="skyblue2",border = "0").pack()
                tkinter.Label(window2,text = "NAME OF COUNTRIES:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
                ent = tkinter.Entry(window2,bd  = "2",bg = "skyblue2")
                ent.place(x = 150 , y = 118)
                def show():
                    h = str(ent.get())
                    a = ""
                    b = []
                    for i in range(0,len(h)):
                        if(h[i]!=","):
                            a = a + h[i]
                        elif (h[i]==','):
                            b.append(a)
                            a = ""
                    for i1 in b: 
                        j = 0  
                        while j < sheet1.nrows:
                            i = 5
                            if sheet.cell_value(j,2).upper() == i1.upper():
                                #to store the death/confirmed cases for every day
                                a1 = []
                                #to store the number of days
                                b1  = []
                                while i < sheet.ncols:
                                    if float(sheet1.cell_value(j,i))>0:
                                        a1.append(float(sheet.cell_value(j,i))/float(sheet1.cell_value(j,i)))
                                    else:
                                        a1.append(0)
                                    b1.append(int(i-4))
                                    i = i + 1
                                if int(var2.get()) == 1:
                                    #plotting a line graph for the same
                                    plt.plot(b1,a1,label= i1)
                                elif int(var1.get()) ==  1:
                                    plt.bar(b1,a1,label = i1)
                                plt.xlabel('DAY')
                                plt.ylabel('DEATH/CONFIRMED CASES')
                                plt.legend()
                                plt.text(0.08, 0.2, i)
                            j = j + 1
                    plt.title(b)
                    plt.show()  
                tkinter.Button(window2,text = "ENTER",bg = "cyan2",command = show).place( x = 260 , y = 200)
                window2.mainloop()
            tkinter.Button(window,text = "ENTER",bg = "cyan2",command = showwwac).place( x = 260 , y = 200)
            window.mainloop()
        tkinter.Button(window2,text = "CURVE OF DEATH RATE",bg = "mediumpurple1",command = deathc).place(x = 240, y = 40)
        def refresh():
            os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx")
            workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx")
            workbook.close()
            url = "https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_recovered_global.csv"
            response = requests.get(url)
            html = response.content
            soup = bs4.BeautifulSoup(html, "lxml")
            tables = soup.findAll("table")
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            table = tables[0]
            for row in table.findAll('tr')[1:]:
                list_of_cells = []
                for cell in row.findAll('td'):
                    text = cell.text.replace('&nbsp;', '')
                    list_of_cells.append(text)
                ws.append(list_of_cells) 
            wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx')
            wb.close()
            os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
            workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
            workbook.close()
            url = "https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_deaths_global.csv"
            response = requests.get(url)
            html = response.content
            soup = bs4.BeautifulSoup(html, "lxml")
            tables = soup.findAll("table")
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            table = tables[0]
            for row in table.findAll('tr')[1:]:
                list_of_cells = []
                for cell in row.findAll('td'):
                    text = cell.text.replace('&nbsp;', '')
                    list_of_cells.append(text)
                ws.append(list_of_cells) 
            wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx')
            wb.close()
            os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
            workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
            workbook.close()
            url = "https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_confirmed_global.csv"
            response = requests.get(url)
            html = response.content
            soup = bs4.BeautifulSoup(html, "lxml")
            tables = soup.findAll("table")
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            table = tables[0]
            for row in table.findAll('tr')[1:]:
                list_of_cells = []
                for cell in row.findAll('td'):
                    text = cell.text.replace('&nbsp;', '')
                    list_of_cells.append(text)
                ws.append(list_of_cells) 
            wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx')
            wb.close()
            os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx")
            workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx")
            workbook.close()
            url = "https://www.worldometers.info/coronavirus"
            response = requests.get(url)
            html = response.content
            soup = bs4.BeautifulSoup(html, "lxml")
            tables = soup.findAll("table")
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            table = tables[0]
            for row in table.findAll('tr')[1:]:
                list_of_cells = []
                for cell in row.findAll('td'):
                    text = cell.text.replace('&nbsp;', '')
                    text = cell.text.replace(',', '')
                    list_of_cells.append(text)
                ws.append(list_of_cells) 
            wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx')
            wb.close()
        tkinter.Button(window2,text = "REFRESH THE DATA",bg = "hotpink",command = refresh).place(x = 20 , y = 220)
        #to show the trends of Deaths
        def coc():
            window2.destroy()
            window = tkinter.Tk()
            window.resizable(0,0)
            window.geometry("400x250")
            window.title("COVID-19 TRACKER")
            canvas = tkinter.Canvas(window,width = 400 , height = 250)
            image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
            canvas.create_image( 199, 124 ,image = image)
            canvas.place(x = 0, y = 0)
            var2 = IntVar()
            var3 = IntVar()
            tkinter.Checkbutton(window,text = "BAR GRAPH",variable = var2,border = "0",bg = "skyblue2",onvalue = 1,offvalue = 0).place(x = 20, y = 148)
            tkinter.Checkbutton(window,text = "LINE GRAPH",variable = var3,border = "0",bg = "skyblue2",onvalue = 1,offvalue = 0).place(x = 150 , y = 148)
            tkinter.Label(window,text = "NUMBER OF COUNTRIES:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
            ent = tkinter.Entry(window,bd  = "2",bg = "skyblue2")
            ent.place(x = 160 , y = 118)
            tkinter.Label(window,text = "*WE RECOMMEND YOU TO SEE BAR GRAPH FOR ONE COUNTRY",bg = "skyblue2",border = "0").place(x = 20 , y = 180)
            def showwwac():
                loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                window.destroy()
                window2 = tkinter.Tk()
                window2.resizable(0,0)
                window2.geometry("400x250")
                window2.title("COVID-19 TRACKER")
                canvas1 = tkinter.Canvas(window2,width = 400 , height = 250)
                image1 = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
                canvas1.create_image( 199, 124 ,image = image1)
                canvas1.place(x = 0, y = 0)
                tkinter.Label(window2,text = "TYPE THE COUNTRIES NAME WITH A COMMA IN END", bg ="skyblue2",border = "0").pack()
                tkinter.Label(window2,text = "NAME OF COUNTRIES:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
                ent = tkinter.Entry(window2,bd  = "2",bg = "skyblue2")
                ent.place(x = 150 , y = 118)
                def show():
                    h = str(ent.get())
                    a = ""
                    b = []
                    for i in range(0,len(h)):
                        if(h[i]!=","):
                            a = a + h[i]
                        elif (h[i]==','):
                            b.append(a)
                            a = ""
                    for i1 in b: 
                        j = 0  
                        while j < sheet.nrows:
                            i = 6
                            if sheet.cell_value(j,2).upper() == i1.upper():
                                #to store the death/confirmed cases for every day
                                a1 = []
                                #to store the number of days
                                b1  = []
                                while i < sheet.ncols:
                                    a1.append(int(sheet.cell_value(j,i)))
                                    b1.append(int(i-4))
                                    i = i + 1
                                if int(var3.get()) == 1:
                                    #plotting a line graph for the same
                                    plt.plot(b1,a1,label= i1)
                                elif int(var2.get()) ==  1:
                                    plt.bar(b1,a1,label = i1)
                                plt.xlabel('DAY')
                                plt.ylabel('DEATHS')
                                plt.legend()
                            j = j + 1
                    plt.title(b)
                    plt.show()  
                tkinter.Button(window2,text = "ENTER",bg = "cyan2",command = show).place( x = 260 , y = 200)
                window2.mainloop()
            tkinter.Button(window,text = "ENTER",bg = "cyan2",command = showwwac).place( x = 260 , y = 200)
            window.mainloop()
        tkinter.Button(window2,text = "CURVE OF DEATHS PER DAY",bg = "maroon",command = coc).place(x = 230 , y = 80)
        #to show the trends of confirmed cases 
        def coc1():
            window2.destroy()
            window = tkinter.Tk()
            window.resizable(0,0)
            window.geometry("400x250")
            window.title("COVID-19 TRACKER")
            canvas = tkinter.Canvas(window,width = 400 , height = 250)
            image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
            canvas.create_image( 199, 124 ,image = image)
            canvas.place(x = 0, y = 0)
            var2 = IntVar()
            var3 = IntVar()
            tkinter.Checkbutton(window,text = "BAR GRAPH",variable = var2,border = "0",bg = "skyblue2",onvalue = 1, offvalue = 0).place(x = 20, y = 148)
            tkinter.Checkbutton(window,text = "LINE GRAPH",variable = var3,border = "0",bg = "skyblue2",onvalue = 1,offvalue = 0).place(x = 150 , y = 148)
            tkinter.Label(window,text = "*WE RECOMMEND YOU TO SEE BAR GRAPH FOR ONE COUNTRY",bg = "skyblue2",border = "0").place(x = 20 , y = 180)
            tkinter.Label(window,text = "NUMBER OF COUNTRIES:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
            ent = tkinter.Entry(window,bd  = "2",bg = "skyblue2")
            ent.place(x = 160 , y = 118)
            def showwwac():
                loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                window.destroy()
                window2 = tkinter.Tk()
                window2.resizable(0,0)
                window2.geometry("400x250")
                window2.title("COVID-19 TRACKER")
                canvas1 = tkinter.Canvas(window2,width = 400 , height = 250)
                image1 = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
                canvas1.create_image( 199, 124 ,image = image1)
                canvas1.place(x = 0, y = 0)
                tkinter.Label(window2,text = "TYPE THE COUNTRIES NAME WITH A COMMA IN END", bg ="skyblue2",border = "0").pack()
                tkinter.Label(window2,text = "NAME OF COUNTRIES:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
                ent = tkinter.Entry(window2,bd  = "2",bg = "skyblue2")
                ent.place(x = 160 , y = 118)
                def show():
                    h = str(ent.get())
                    a = ""
                    b = []
                    for i in range(0,len(h)):
                        if(h[i]!=","):
                            a = a + h[i]
                        elif (h[i]==','):
                            b.append(a)
                            a = ""
                    for i1 in b: 
                        j = 0  
                        while j < sheet.nrows:
                            i = 6
                            if sheet.cell_value(j,2).upper() == i1.upper():
                                #to store the death/confirmed cases for every day
                                a1 = []
                                #to store the number of days
                                b1  = []
                                while i < sheet.ncols:
                                    a1.append(int(sheet.cell_value(j,i)))
                                    b1.append(int(i-4))
                                    i = i + 1
                                if int(var3.get()) == 1:
                                    #plotting a line graph for the same
                                    plt.plot(b1,a1,label= i1)
                                elif int(var2.get()) ==  1:
                                    plt.bar(b1,a1,label = i1)
                                plt.xlabel('DAY')
                                plt.ylabel('CONFIRMED')
                                plt.legend()
                            j = j + 1
                    plt.title(b)
                    plt.show()  
                tkinter.Button(window2,text = "ENTER",bg = "cyan2",command = show).place( x = 260 , y = 200)
                window2.mainloop()
            tkinter.Button(window,text = "ENTER",bg = "cyan2",command = showwwac).place( x = 260 , y = 200)
            window.mainloop()
        tkinter.Button(window2,text = "CURVE OF CONFIRMED CASES PER DAY",bg = "yellow",command = coc1).place(x = 180 , y = 120)
        #to show the trends of recovered
        def coc2():
            window2.destroy()
            window = tkinter.Tk()
            window.resizable(0,0)
            window.geometry("400x250")
            window.title("COVID-19 TRACKER")
            canvas = tkinter.Canvas(window,width = 400 , height = 250)
            image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
            canvas.create_image( 199, 124 ,image = image)
            canvas.place(x = 0, y = 0)
            var2 = IntVar()
            var3 = IntVar()
            tkinter.Checkbutton(window,text = "BAR GRAPH",variable = var2,border = "0",bg = "skyblue2",onvalue = 1,offvalue = 0).place(x = 20, y = 148)
            tkinter.Checkbutton(window,text = "LINE GRAPH",variable = var3,border = "0",bg = "skyblue2",onvalue = 1,offvalue = 0).place(x = 150 , y = 148)
            tkinter.Label(window,text = "*WE RECOMMEND YOU TO SEE BAR GRAPH FOR ONE COUNTRY",bg = "skyblue2",border = "0").place(x = 20 , y = 180)
            tkinter.Label(window,text = "NUMBER OF COUNTRIES:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
            ent = tkinter.Entry(window,bd  = "2",bg = "skyblue2")
            ent.place(x = 160 , y = 118)
            def showwwac():
                loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                window.destroy()
                window2 = tkinter.Tk()
                window2.resizable(0,0)
                window2.geometry("400x250")
                window2.title("COVID-19 TRACKER")
                canvas1 = tkinter.Canvas(window2,width = 400 , height = 250)
                image1 = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
                canvas1.create_image( 199, 124 ,image = image1)
                canvas1.place(x = 0, y = 0)
                tkinter.Label(window2,text = "TYPE THE COUNTRIES NAME WITH A COMMA IN END", bg ="skyblue2",border = "0").pack()
                tkinter.Label(window2,text = "NAME OF COUNTRIES:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
                ent = tkinter.Entry(window2,bd  = "2",bg = "skyblue2")
                ent.place(x = 150 , y = 118)
                def show():
                    h = str(ent.get())
                    a = ""
                    b = []
                    for i in range(0,len(h)):
                        if(h[i]!=","):
                            a = a + h[i]
                        elif (h[i]==','):
                            b.append(a)
                            a = ""
                    for i1 in b: 
                        j = 0  
                        while j < sheet.nrows:
                            i = 6
                            if sheet.cell_value(j,2).upper() == i1.upper():
                                #to store the death/confirmed cases for every day
                                a1 = []
                                #to store the number of days
                                b1  = []
                                while i < sheet.ncols:
                                    if(int(sheet1.cell_value(j,i))!=0):
                                        a1.append(float(sheet.cell_value(j,i)))
                                    else:
                                        a1.append(0)
                                    b1.append(int(i-4))
                                    i = i + 1
                                if int(var3.get()) == 1:
                                    #plotting a line graph for the same
                                    plt.plot(b1,a1,label= i1)
                                elif int(var2.get()) ==  1:
                                    plt.bar(b1,a1,label = i1)
                                plt.xlabel('DAY')
                                plt.ylabel('RECOVERED')
                                plt.legend()
                            j = j + 1
                    plt.title(b)
                    plt.show()  
                tkinter.Button(window2,text = "ENTER",bg = "cyan2",command = show).place( x = 260 , y = 200)
                window2.mainloop()
            tkinter.Button(window,text = "ENTER",bg = "cyan2",command = showwwac).place( x = 260 , y = 200)
            window.mainloop()
        tkinter.Button(window2,text = "CURVE OF RECOVERED CASES PER DAY",bg = "green",command = coc2).place(x = 180 , y = 160)
        window2.resizable(0,0)
        window2.mainloop()
    tkinter.Button(window,text = "GRAPHICAL DATA",bg = "magenta",command = graphs,border = "0").place(x = 240 , y = 40)
    #showing all the tabular data just open to know more
    def tabular():
        window.destroy()
        window2 = tkinter.Tk()
        window2.title("COVID-19 TRACKER")
        window2.geometry("400x250")
        canvas = tkinter.Canvas(window2,width = 400 , height = 250)
        image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
        canvas.create_image( 199, 124 ,image = image)
        canvas.place(x = 0, y = 0)
        #this function here is to show the condition of the country
        
        def country():
            window2.destroy()
            window = tkinter.Tk()
            window.title("COVID-19 TRACKER")
            window.geometry("400x250")
            #adding the photo
            canvas = tkinter.Canvas(window,width = 400 , height = 250)
            image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
            canvas.create_image( 199, 124 ,image = image)
            canvas.place(x = 0, y = 0)
            ent = tkinter.Entry(window,bd = 2,bg ="dark khaki")
            ent.place(x = 100 , y = 140)
            tkinter.Label(window,text = "COUNTRY:",bg ="dark khaki",border ="0").place(x = 35,y = 142)
            def showt():
                h = ent.get()
                window.destroy()
                window3 = tkinter.Tk()
                window3.title("COVID 19 TRACKER")
                body = tkinter.ttk.Frame(window3)  
                body.pack(fill=tkinter.BOTH, expand=True)
                scrollable_frame = Scrollable(body, width=32)
                loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                tkinter.Label(scrollable_frame,text = "COUNTRY").grid(row = 1, column = 0, pady = 00, ipadx = 40)
                tkinter.Label(scrollable_frame,text = "TOTAL CASES").grid(row = 1, column = 1, pady = 00, ipadx = 20)
                tkinter.Label(scrollable_frame,text = "NEW CASES").grid(row = 1, column = 2, pady = 00, ipadx = 20)
                tkinter.Label(scrollable_frame,text = "TOTAL DEATH").grid(row = 1, column = 3, pady = 00, ipadx = 20)
                tkinter.Label(scrollable_frame,text = "NEW DEATH").grid(row = 1, column = 4, pady = 00, ipadx = 20)
                tkinter.Label(scrollable_frame,text = "TOTAL RECOVERED").grid(row = 1, column = 5, pady = 00, ipadx = 20)
                tkinter.Label(scrollable_frame,text = "ACTIVE CASES").grid(row = 1, column = 6, pady = 00, ipadx = 20)
                tkinter.Label(scrollable_frame,text = "SERIOUS CASES").grid(row = 1, column = 7, pady = 00, ipadx = 20)
                r = 2
                g = h.upper()
                print(g)
                for i in range(0,sheet.nrows):
                    if(str(sheet.cell_value(i,0)).upper() == g):
                        if(sheet.cell_value(i,3)==' '):
                            b = 0
                        else:
                            b = int(sheet.cell_value(i,3))
                        #here we are counting the death rate if it is greater than 2% the country is not suitable for us to go for next 2 years 
                        if(b/int(sheet.cell_value(i,1))*100>=2):
                            tkinter.Label(scrollable_frame,text = "YOUR COUNTRY IS NOT SUITABLE FOR PROJECT",border = "0",bg = "red",font = "ARIALBLACK").grid(row = 0, column = 0, pady = 10, ipadx = 20)
                            tkinter.Label(scrollable_frame,text = sheet.cell_value(i,0),border = "0",bg = "red").grid(row = r, column = 0, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,1)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1),border = "0",bg = "red").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,2)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,2),border = "0",bg = "red").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,3)):
                                tkinter.Label(scrollable_frame,text = 0,border = "0",bg = "red").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,3),border = "0",bg = "red").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,4)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,4),border = "0",bg = "red").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,5)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,5),border = "0",bg = "red").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,6)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,6),border = "0",bg = "red").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,7)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 7, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,7),border = "0",bg = "red").grid(row = r, column = 7, pady = 00, ipadx = 20)
                        
                        #here we are counting the death rate if it is less than 2%  and greater than 1.5 the country is somewhat suitable for us to go for next 2 years
                        if(int(sheet.cell_value(i,3))/int(sheet.cell_value(i,1))*100<2) and (int(sheet.cell_value(i,3))/int(sheet.cell_value(i,1))*100>=1.5):
                            tkinter.Label(scrollable_frame,text = "YOUR COUNTRY IS SOMEWHAT SUITABLE FOR PROJECT",border = "0",bg = "gold",font = "ARIALBLACK").grid(row = 0, column = 0, pady = 10, ipadx = 20)
                            tkinter.Label(scrollable_frame,text = sheet.cell_value(i,0),border = "0",bg = "gold").grid(row = r, column = 0, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,1)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1),border = "0",bg = "gold").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,2)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,2),border = "0",bg = "gold").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,3)):
                                tkinter.Label(scrollable_frame,text = 0,border = "0",bg = "gold").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,3),border = "0",bg = "gold").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,4)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,4),border = "0",bg = "gold").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,5)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,5),border = "0",bg = "gold").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,6)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,6),border = "0",bg = "gold").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,7)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 7, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,7),border = "0",bg = "gold").grid(row = r, column = 7, pady = 00, ipadx = 20)                
                        
                        #here we are counting the death rate if it is less than 1.5% the country is suitable for us to go for next 2 years
                        if(int(sheet.cell_value(i,3))/int(sheet.cell_value(i,1))*100<1.5):
                            tkinter.Label(scrollable_frame,text = "YOUR COUNTRY IS SUITABLE FOR PROJECT",border = "0",bg = "green",font = "ARIALBLACK").grid(row = 0, column = 0, pady = 10, ipadx = 20)
                            tkinter.Label(scrollable_frame,text = sheet.cell_value(i,0),border = "0",bg = "green").grid(row = r, column = 0, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,1)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1),border = "0",bg = "green").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,2)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,2),border = "0",bg = "green").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,3)):
                                tkinter.Label(scrollable_frame,text = 0,border = "0",bg = "green").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,3),border = "0",bg = "green").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,4)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,4),border = "0",bg = "green").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,5)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,5),border = "0",bg = "green").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,6)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,6),border = "0",bg = "green").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,7)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 7, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,7),border = "0",bg = "green").grid(row = r, column = 7, pady = 00, ipadx = 20)                
                r =  3
                for i in range(8,sheet.nrows-8):
                    if not (sheet.cell_value(i,0).upper() == h):
                        if(sheet.cell_value(i,3)==' '):
                            b = 0
                        else:
                            b = int(sheet.cell_value(i,3))
                        if(b/int(sheet.cell_value(i,1))*100>=2):
                            tkinter.Label(scrollable_frame,text = sheet.cell_value(i,0),border = "0",bg = "red").grid(row = r, column = 0, pady = 5, ipadx = 20)
                            if(not sheet.cell_value(i,1)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1),border = "0",bg = "red").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,2)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,2),border = "0",bg = "red").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,3)):
                                tkinter.Label(scrollable_frame,text = 0,border = "0",bg = "red").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,3),border = "0",bg = "red").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,4)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,4),border = "0",bg = "red").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,5)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,5),border = "0",bg = "red").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,6)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,6),border = "0",bg = "red").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,7)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "red").grid(row = r, column = 7, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,7),border = "0",bg = "red").grid(row = r, column = 7, pady = 00, ipadx = 20)
                    r =  r + 1
                for i in range(8,sheet.nrows-8):
                    if(sheet.cell_value(i,3)==' '):
                            b = 0
                    else:
                            b = int(sheet.cell_value(i,3))
                    if not (sheet.cell_value(i,0).upper() == g):
                        if((b/int(sheet.cell_value(i,1))*100<2) and (b/int(sheet.cell_value(i,1))*100>=1.5)):
                            tkinter.Label(scrollable_frame,text = sheet.cell_value(i,0),border = "0",bg = "gold").grid(row = r, column = 0, pady = 5, ipadx = 20)
                            if(not sheet.cell_value(i,1)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1),border = "0",bg = "gold").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,2)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,2),border = "0",bg = "gold").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,3)):
                                tkinter.Label(scrollable_frame,text = 0,border = "0",bg = "gold").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,3),border = "0",bg = "gold").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,4)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,4),border = "0",bg = "gold").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,5)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,5),border = "0",bg = "gold").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,6)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,6),border = "0",bg = "gold").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,7)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "gold").grid(row = r, column = 7, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,7),border = "0",bg = "gold").grid(row = r, column = 7, pady = 00, ipadx = 20)
                    r =  r + 1
                for i in range(8,sheet.nrows-8):
                    if(sheet.cell_value(i,3)==' '):
                            b = 0
                    else:
                            b = int(sheet.cell_value(i,3))
                    if not (sheet.cell_value(i,0).upper() == g):
                        if(b/int(sheet.cell_value(i,1))*100<1.5):
                            tkinter.Label(scrollable_frame,text = sheet.cell_value(i,0),border = "0",bg = "green").grid(row = r, column = 0, pady = 5, ipadx = 20)
                            if(not sheet.cell_value(i,1)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1),border = "0",bg = "green").grid(row = r, column = 1, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,2)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,2),border = "0",bg = "green").grid(row = r, column = 2, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,3)):
                                tkinter.Label(scrollable_frame,text = 0,border = "0",bg = "green").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,3),border = "0",bg = "green").grid(row = r, column = 3, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,4)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,4),border = "0",bg = "green").grid(row = r, column = 4, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,5)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,5),border = "0",bg = "green").grid(row = r, column = 5, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,6)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,6),border = "0",bg = "green").grid(row = r, column = 6, pady = 00, ipadx = 20)
                            if(not sheet.cell_value(i,7)):
                                tkinter.Label(scrollable_frame,text = "0",border = "0",bg = "green").grid(row = r, column = 7, pady = 00, ipadx = 20)
                            else:
                                tkinter.Label(scrollable_frame,text = sheet.cell_value(i,7),border = "0",bg = "green").grid(row = r, column = 7, pady = 00, ipadx = 20)                
                    r = r + 1                
                scrollable_frame.update()
                window3.mainloop()
            tkinter.Button(window,text = "ENTER",bg = "tan1",command  = showt).place(x = 100 , y = 180)
            window.resizable(0,0)
            window.mainloop()
        tkinter.Button(window2,text = "CONDITION OF COUNTRIES",bg = "purple1",command  = country).place(x = 230 , y = 40)
        #this function is to refresh the data
        def refresh():
            #removing the file
            os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx")
            #creating the new file
            workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx")
            workbook.close()
            #opening the url
            url = "https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_recovered_global.csv"
            response = requests.get(url)
            html = response.content
            #opening the html format of web page
            soup = bs4.BeautifulSoup(html, "lxml")
            #founding all the tables in the web page
            tables = soup.findAll("table")
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            #opening the first table
            table = tables[0]
            #<tr> tag is the tag in html for rows so we are iterating through the rows
            for row in table.findAll('tr')[1:]:
                list_of_cells = []
                #<td> tag is for cell data 
                for cell in row.findAll('td'):
                    #here we are deleting all the tags other than the data in the cell
                    text = cell.text.replace('&nbsp;', '')
                    #storing it in the list_of_cells list
                    list_of_cells.append(text)
                ws.append(list_of_cells)
            #saving the data in the xlsx file 
            wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx')
            wb.close()
            #same with the different files
            os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
            workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
            workbook.close()
            url = "https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_deaths_global.csv"
            response = requests.get(url)
            html = response.content
            soup = bs4.BeautifulSoup(html, "lxml")
            tables = soup.findAll("table")
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            table = tables[0]
            for row in table.findAll('tr')[1:]:
                list_of_cells = []
                for cell in row.findAll('td'):
                    text = cell.text.replace('&nbsp;', '')
                    list_of_cells.append(text)
                ws.append(list_of_cells) 
            wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx')
            wb.close()
            os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
            workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
            workbook.close()
            url = "https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_confirmed_global.csv"
            response = requests.get(url)
            html = response.content
            soup = bs4.BeautifulSoup(html, "lxml")
            tables = soup.findAll("table")
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            table = tables[0]
            for row in table.findAll('tr')[1:]:
                list_of_cells = []
                for cell in row.findAll('td'):
                    text = cell.text.replace('&nbsp;', '')
                    list_of_cells.append(text)
                ws.append(list_of_cells) 
            wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx')
            wb.close()
            os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx")
            workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx")
            workbook.close()
            url = "https://www.worldometers.info/coronavirus"
            response = requests.get(url)
            html = response.content
            soup = bs4.BeautifulSoup(html, "lxml")
            tables = soup.findAll("table")
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            table = tables[0]
            for row in table.findAll('tr')[1:]:
                list_of_cells = []
                for cell in row.findAll('td'):
                    text = cell.text.replace('&nbsp;', '')
                    text = cell.text.replace(',', '')
                    list_of_cells.append(text)
                ws.append(list_of_cells) 
            wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Countries.xlsx')
            wb.close()
        tkinter.Button(window2,text = "REFRESH THE DATA",bg = "hotpink",command = refresh).place(x = 20 , y = 220)
        #this function here is to calculate the average confirmed cases per week
        def wwac():
            window2.destroy()
            window = tkinter.Tk()
            window.resizable(0,0)
            window.geometry("400x250")
            window.title("COVID-19 TRACKER")
            #these are steps to add the photo in the frame
            canvas = tkinter.Canvas(window,width = 400 , height = 250)
            image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
            canvas.create_image( 199, 124 ,image = image)
            canvas.place(x = 0, y = 0)
            tkinter.Label(window,text = "COUNTRY NAME:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
            ent = tkinter.Entry(window,bd  = "2",bg = "skyblue2")
            ent.place(x = 120 , y = 118)
            #this function here is to calculate the average confirmed cases per week
            def showwwac():
                #file where the data of the confirmed cases is stored
                loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Confirmed.xlsx")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                #calculating the number of weeks
                no_of_weeks = sheet.ncols - 5
                no_of_weeks = int(no_of_weeks / 7)
                h  = ent.get()
                window.destroy()
                window2 = tkinter.Tk()
                window2.geometry("400x250")
                window2.title("COVID-19 TRACKER")
                h = h.upper()
                #adding the scroll bar on the window
                body = tkinter.ttk.Frame(window2)  
                body.pack(fill=tkinter.BOTH, expand=True)
                scrollable_frame = Scrollable(body, width=32)
                tkinter.Label(scrollable_frame,text = h,bg = "yellow1",border = "0",font = "ARIAL").grid(row = 0, column = 5, pady = 5 , padx = 10)
                tkinter.Label(scrollable_frame,text = "WEEK NUMBER",bg = "yellow1",border = "0").grid(row = 1, column = 0, pady = 5 , padx = 10)
                tkinter.Label(scrollable_frame,text = "AVERAGE CONFIRMED CASE PER WEEK",bg = "yellow1",border = "0").grid(row = 1, column = 3, pady = 5 , padx = 10)
                y = 2
                for i in range(0 , sheet.nrows):
                    #checking whether the given country is there or not
                    if (sheet.cell_value(i,2).upper() == h):
                        tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1).upper(),bg = "yellow1").grid(row = y  , column = 5 , pady = 5 , padx = 10)
                        y = y + 1
                        j = 0
                        q = 5
                        #storing the data in form of the
                        a = []
                        while j < no_of_weeks:
                            sum = 0.0
                            sum = sum + float(sheet.cell_value(i,q+7))-float(sheet.cell_value(i,q))
                            q = q + 7
                            j = j + 1
                            sum = int(sum/7.0)
                            sum = str(sum)
                            a.append(sum)
                        o = 1
                        for k in range(0 , no_of_weeks):
                            week = "WEEK " + str(o) + " :"
                            o = o + 1
                            tkinter.Label(scrollable_frame,text = week,bg = "yellow2",border = "0").grid(row  = y , column = 0, pady = 5 , padx = 10)
                            tkinter.Label(scrollable_frame,text = a[k],bg = "yellow2",border = "0").grid(row  = y , column = 3, pady = 5 , padx = 10)
                            y = y + 1
                scrollable_frame.update()   
                window2.mainloop()
            tkinter.Button(window,text = "ENTER",bg = "cyan2",command = showwwac).place( x = 260 , y = 200)
            window.mainloop()
        tkinter.Button(window2,text = "WEEK WISE AVERAGE CONFIRMED",bg = "yellow",command = wwac).place(x = 206 , y = 160)
        #this function here is to calculate the average death cases per week
        def wwad():
            window2.destroy()
            window = tkinter.Tk()
            window.resizable(0,0)
            window.geometry("400x250")
            window.title("COVID-19 TRACKER")
            #these are steps to add the photo in the frame
            canvas = tkinter.Canvas(window,width = 400 , height = 250)
            image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
            canvas.create_image( 199, 124 ,image = image)
            canvas.place(x = 0, y = 0)
            tkinter.Label(window,text = "COUNTRY NAME:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
            ent = tkinter.Entry(window,bd  = "2",bg = "skyblue2")
            ent.place(x = 120 , y = 118)
            #the function mainly count the average number of dead paitents per week and store them in a list and then show them
            def showwwac():
                #file which contains the data of per day Dead paitents
                loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Deaths.xlsx")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                #here as our data is in time series and day wise so no of weeks will be calculated as no of days which is number of columns / 7
                no_of_weeks = sheet.ncols - 5
                no_of_weeks = int(no_of_weeks / 7)
                h  = ent.get()
                window.destroy()
                window2 = tkinter.Tk()
                window2.geometry("400x250")
                window2.title("COVID-19 TRACKER")
                h = h.upper()
                #these are steps to add the photo in the frame
                body = tkinter.ttk.Frame(window2)  
                body.pack(fill=tkinter.BOTH, expand=True)
                scrollable_frame = Scrollable(body, width=32)
                tkinter.Label(scrollable_frame,text = h,bg = "red",border = "0",font = "ARIAL").grid(row = 0, column = 5, pady = 5 , padx = 10)
                tkinter.Label(scrollable_frame,text = "WEEK NUMBER",bg = "red",border = "0").grid(row = 1, column = 0, pady = 5 , padx = 10)
                tkinter.Label(scrollable_frame,text = "AVERAGE DEATH CASE PER WEEK",bg = "red",border = "0").grid(row = 1, column = 3, pady = 5 , padx = 10)
                y = 2
                for i in range(0 , sheet.nrows):
                    #here we are checking the cell_value of the column number 3 as it is the country name 
                    if (sheet.cell_value(i,2).upper() == h):
                        tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1).upper(),bg = "red").grid(row = y  , column = 5 , pady = 5 , padx = 10)
                        y = y + 1
                        j = 0
                        q = 5
                        #this is the list to dead cases for that week
                        a = []
                        while j < no_of_weeks:
                            sum = 0.0
                            #calculating as dead cases till nth week - recovered cases till (n-1)th week and storing it in a
                            sum = sum + float(sheet.cell_value(i,q+7))-float(sheet.cell_value(i,q))
                            q = q + 7
                            j = j + 1
                            sum = int(sum/7.0)
                            sum = str(sum)
                            a.append(sum)
                        o = 1
                        for k in range(0 , no_of_weeks):
                            week = "WEEK " + str(o) + " :"
                            o = o + 1
                            tkinter.Label(scrollable_frame,text = week,bg = "red",border = "0").grid(row  = y , column = 0, pady = 5 , padx = 10)
                            tkinter.Label(scrollable_frame,text = a[k],bg = "red",border = "0").grid(row  = y , column = 3, pady = 5 , padx = 10)
                            y = y + 1
                scrollable_frame.update()   
                window2.mainloop()
            tkinter.Button(window,text = "ENTER",bg = "cyan2",command = showwwac).place( x = 260 , y = 200)
            window.mainloop()
        tkinter.Button(window2,text = "WEEK WISE AVERAGE DEATHS",bg = "maroon",command = wwad).place(x = 222 , y = 80 )
        #this function here is to calculate the average recovered cases per week
        def wwar():
            window2.destroy()
            window = tkinter.Tk()
            window.resizable(0,0)
            window.geometry("400x250")
            window.title("COVID-19 TRACKER")
            #these are steps to add the photo in the frame
            canvas = tkinter.Canvas(window,width = 400 , height = 250)
            image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
            canvas.create_image( 199, 124 ,image = image)
            canvas.place(x = 0, y = 0)
            tkinter.Label(window,text = "COUNTRY NAME:",bg = "skyblue2",border = "0").place(x = 20, y = 120)
            ent = tkinter.Entry(window,bd  = "2",bg = "skyblue2")
            ent.place(x = 120 , y = 118)
            #the function mainly count the average number of recovered paitents per week and store them in a list and then show them
            def showwwac():
                #file which contains the data of per day recovered paitents
                loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/Recovered.xlsx")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                #here as our data is in time series and day wise so no of weeks will be calculated as no of days which is number of columns / 7
                no_of_weeks = sheet.ncols - 5
                no_of_weeks = int(no_of_weeks / 7)
                h  = ent.get()
                window.destroy()
                window2 = tkinter.Tk()
                window2.geometry("400x250")
                window2.title("COVID-19 TRACKER")
                h = h.upper()
                #here this body and scrollable_frame are used to add the scrollbar in the frame
                body = tkinter.ttk.Frame(window2)  
                body.pack(fill=tkinter.BOTH, expand=True)
                scrollable_frame = Scrollable(body, width=32)
                tkinter.Label(scrollable_frame,text = h,bg = "limegreen",border = "0",font = "ARIAL").grid(row = 0, column = 5, pady = 5 , padx = 10)
                tkinter.Label(scrollable_frame,text = "WEEK NUMBER",bg = "limegreen",border = "0").grid(row = 1, column = 0, pady = 5 , padx = 10)
                tkinter.Label(scrollable_frame,text = "AVERAGE RECOVERED CASE PER WEEK",bg = "limegreen",border = "0").grid(row = 1, column = 3, pady = 5 , padx = 10)
                y = 2
                for i in range(0 , sheet.nrows):
                    #here we are checking the cell_value of the column number 3 as it is the country name 
                    if (sheet.cell_value(i,2).upper() == h):
                        tkinter.Label(scrollable_frame,text = sheet.cell_value(i,1).upper(),bg = "limegreen").grid(row = y  , column = 5 , pady = 5 , padx = 10)
                        y = y + 1
                        j = 0
                        q = 5
                        #this is the list to store recovered cases for that week
                        a = []
                        while j < no_of_weeks:
                            sum = 0.0
                            #calculating as recoverd cases till nth week - recovered cases till (n-1)th week and storing it in a 
                            sum = sum + float(sheet.cell_value(i,q+7))-float(sheet.cell_value(i,q))
                            q = q + 7
                            j = j + 1
                            sum = int(sum/7.0)
                            sum = str(sum)
                            a.append(sum)
                        o = 1
                        for k in range(0 , no_of_weeks):
                            week = "WEEK " + str(o) + " :"
                            o = o + 1
                            tkinter.Label(scrollable_frame,text = week,bg = "limegreen",border = "0").grid(row  = y , column = 0, pady = 5 , padx = 10)
                            tkinter.Label(scrollable_frame,text = a[k],bg = "limegreen",border = "0").grid(row  = y , column = 3, pady = 5 , padx = 10)
                            y = y + 1
                scrollable_frame.update()   
                window2.mainloop()
            tkinter.Button(window,text = "ENTER",bg = "cyan2",command = showwwac).place( x = 260 , y = 200)
            window.mainloop()
        tkinter.Button(window2,text = "WEEK WISE AVERAGE RECOVERED",bg = "limegreen",command = wwar).place(x = 208 , y = 120 )
        window2.resizable(0,0)
        window2.mainloop()
    tkinter.Button(window,text = "TABULAR DATA",bg = "cyan",command = tabular,border = "0").place(x = 247, y = 120)
    #it is a simple interface which contains the set of question about the symptoms of COVID-19 which will help us to make whether the user is suffering from 
    #COVID-19 or not
    def ask():
        window.destroy()
        window2 = tkinter.Tk()
        window2.title("COVID-19 TRACKER")
        window2.geometry("400x250")
        #these are there to store the response of checkbox
        var1 = IntVar()
        var2 = IntVar()
        var3 = IntVar()
        var4 = IntVar()
        var5 = IntVar()
        var6 = IntVar()
        var7 = IntVar()
        var8 = IntVar()
        var9 = IntVar()
        var10 = IntVar()
        count = 0
        tkinter.Label(window2,text = "1.  Do you have cough?").place(x = 10,y = 20)
        tkinter.Checkbutton(window2, text = "YES",variable = var1,command = count,onvalue = 1, offvalue = 0).place(x = 20, y = 40)
        tkinter.Checkbutton(window2, text = "NO",variable = var2,command = count,onvalue = 1, offvalue = 0).place(x = 70, y = 40)
        tkinter.Label(window2,text = "2.  Do you have fever?").place(x = 10,y = 60)
        tkinter.Checkbutton(window2, text = "YES",variable = var3,command = count,onvalue = 1, offvalue = 0).place(x = 20, y = 80)
        tkinter.Checkbutton(window2, text = "NO",variable = var4,command = count,onvalue = 1, offvalue = 0).place(x = 70, y = 80)
        tkinter.Label(window2,text = "3.  Does your sense of smell gone?").place(x = 10,y = 100)
        tkinter.Checkbutton(window2, text = "YES",variable = var5,command = count,onvalue = 1, offvalue = 0).place(x = 20, y = 120)
        tkinter.Checkbutton(window2, text = "NO",variable = var6,command = count,onvalue = 1, offvalue = 0).place(x = 70, y = 120)
        tkinter.Label(window2,text = "4.  Do you have a abroad travel history?").place(x = 10,y = 140)
        tkinter.Checkbutton(window2, text = "YES",variable = var7,command = count,onvalue = 1, offvalue = 0).place(x = 20, y = 160)
        tkinter.Checkbutton(window2, text = "NO",variable = var8,command = count,onvalue = 1, offvalue = 0).place(x = 70, y = 160)
        tkinter.Label(window2,text = "5.  Did you come in contact of COVID-19 patient?").place(x = 10,y = 180)
        tkinter.Checkbutton(window2, text = "YES",variable = var9,command = count,onvalue = 1, offvalue = 0).place(x = 20, y = 200)
        tkinter.Checkbutton(window2, text = "NO",variable = var10,command = count,onvalue = 1, offvalue = 0).place(x = 70, y = 200)
        def county():
            #to count the number of yes
            count = 0
            #here we are counting the number of yes the value will be 1 if it is corrected and 0 if not
            count = int(var1.get()) + int(var3.get()) + int(var5.get()) + int(var7.get()) + int(var9.get())
            window2.destroy()
            window3 = tkinter.Tk()
            window3.title("COVID-19 TRACKER")
            window3.geometry("400x260")
             #there is one more addition in the window which is helpine nunmbers for every country  
            def showlist():
                r = "https://www.countryliving.com/uk/wildlife/countryside/news/a1553/emergency-numbers-in-countries-abroad/"
                new = 1
                #opening the webbrowser
                webbrowser.open(r,new = new)
            #helpline button to redirect the user to the list of emergency numbers of every country
            tkinter.Button(window3,text = "HELPLINE",border = "0",command = showlist,bg = "brown").place(x=0,y=0)
            wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/patiemt.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")
            #if he is having less than 2 symptoms the rate according to WHO is low
            if count < 2:
                tkinter.Label(window3,text = "LOW RISK OF INFECTION",font = "arial",bg = "green").pack()
                image = tkinter.PhotoImage(file = "C:\\Users\\yasharth dubey\\Documents\\Python Scripts\\COVID19 management\\0.png")
                tkinter.Label(window3,image = image).pack()
                ws.cell(row = 2, column = 1).value  = int(ws.cell(row = 2,column = 1).value) + 1
            #if the symptom count is above 2 and below 4 the risk of COVID-19 is medium and he may consider a doctor
            elif count>=2 and count<4:
                window3.geometry("410x260")
                tkinter.Label(window3,text = "MEDIUM RISK OF INFECTION",font = "arial",bg = "yellow").pack()
                image = tkinter.PhotoImage(file = "C:\\Users\\yasharth dubey\\Documents\\Python Scripts\\COVID19 management\\1.png")
                tkinter.Label(window3,image = image).pack() 
                ws.cell(row = 2, column = 2).value  = int(ws.cell(row = 2,column = 2).value) + 1
            #in any other condition the risk is high  
            else:
                tkinter.Label(window3,text = "HIGH RISK OF INFECTION",font = "arial",bg = "maroon").pack()
                image = tkinter.PhotoImage(file = "C:\\Users\\yasharth dubey\\Documents\\Python Scripts\\COVID19 management\\2.png")
                tkinter.Label(window3,image = image).pack()
                ws.cell(row = 2, column = 3).value  = int(ws.cell(row = 2,column = 3).value) + 1
            wb.save("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/patiemt.xlsx")
            wb.close()
            a = []
            b = []
            loc  = ("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/patiemt.xlsx")
            wb = xlrd.open_workbook(loc)
            sheet = wb.sheet_by_index(0)
            a.append(sheet.cell_value(0,0))
            a.append(sheet.cell_value(0,1))
            a.append(sheet.cell_value(0,2))
            b.append(int(sheet.cell_value(1,0)))
            b.append(int(sheet.cell_value(1,1)))
            b.append(int(sheet.cell_value(1,2)))
            plt.bar(a[0],b[0],label = "RISK",color = "green")
            plt.bar(a[1],b[1],label = "RISK",color = "yellow")
            plt.bar(a[2],b[2],label = "RISK",color = "red")
            plt.show()
            window3.resizable(0,0)
            window3.mainloop()
        tkinter.Button(window2,text = "ENTER",border = "0",bg = "orange",command = county).place(x = 180 , y = 220 )
        window2.resizable(0,0)
        window2.mainloop()
    tkinter.Button(window,text = "CHECK YOURSELF",bg = "yellow",command = ask,border = "0").place(x = 240, y = 80)
    #simple introduction about the group members
    def about():
        window.destroy()
        window2 = tkinter.Tk()
        window2.title("COVID-19 TRACKER")
        window2.geometry("400x250")
        canvas = tkinter.Canvas(window2,width = 400 , height = 250)
        image = ImageTk.PhotoImage(Image.open("C:/Users/yasharth dubey/Documents/Python Scripts/COVID19 management/images.png"))
        canvas.create_image( 199, 124 ,image = image)
        canvas.place(x = 0, y = 0)
        def mem1():
            r = "https://www.linkedin.com/in/yasharth-dubey-0434b6155"
            new = 1
            webbrowser.open(r,new = new)
        tkinter.Button(window2,text = "MEMBER 1",bg = "purple1",command = mem1).place(x = 250 , y = 40)
        def mem2():
            r = "https://www.linkedin.com/mwlite/in/vikash-singh-06572a1a7"
            new = 1
            webbrowser.open(r,new = new)
        tkinter.Button(window2,text = "MEMBER 2",bg = "hotpink",command = mem2).place(x = 250 , y = 80)
        def mem3():
            r = "https://www.linkedin.com/in/afeef-hasan-27772b1a7/"
            new = 1
            webbrowser.open(r,new = new)
        tkinter.Button(window2,text = "MEMBER 3",bg = "maroon",command = mem3).place(x = 250 , y = 120)
        def mem4():
            r = "https://www.linkedin.com/in/ashwin-mishra-84a408187"
            new = 1
            webbrowser.open(r,new = new)
        tkinter.Button(window2,text = "MEMBER 4",bg = "yellow",command = mem4).place(x = 250 , y = 160)
        window2.mainloop()
    tkinter.Button(window,text = "ABOUT US",bg = "saddle brown",command = about,border = "0").place(x = 260,y = 160)
    #the fuction redirects to WHO site to make us know better about COVID-19
    def KNOW():
        r = "https://www.who.int/health-topics/coronavirus#tab=tab_1"
        new = 1
        webbrowser.open(r,new = new)
    tkinter.Button(window,text = "KNOW MORE",bg = "seagreen",command = KNOW,border = "0").place(x = 20,y = 200)
    #To exit the Application
    def end():
        exit(0)
    tkinter.Button(window,text = "EXIT",bg = "lime",command = end,border = "0").place(x = 273, y = 200)
    window.resizable(0,0)
    window.mainloop()
#except block
except:
    tkinter.messagebox.showerror(title = "ERROR", message = "PROGRAM TERMINATING")
#finally block
finally:
    tkinter.messagebox.showinfo(title = "COVID-19 TRACKER", message = "THANKS FOR USING OUR GUI \n SEND US FEEDBACK to 457ayan123@gmail.com")
